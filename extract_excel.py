import sys
import os

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.run([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
    import openpyxl

def read_excel(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n=== Sheet: {sheet_name} ===")
        for row in ws.iter_rows(values_only=True):
            if any(v is not None for v in row):
                print(row)
    wb.close()

files = [
    r'C:\Ai\BattleAgent\Data\Skill.xlsx',
    r'C:\Ai\BattleAgent\Data\MonsterGroup.xlsx',
    r'C:\Ai\BattleAgent\Data\Character.xlsx',
    r'C:\Ai\BattleAgent\Data\Stage.xlsx',
]

for f in files:
    print(f"\n\n{'='*60}")
    print(f"FILE: {f}")
    print('='*60)
    read_excel(f)
