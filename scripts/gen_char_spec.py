#!/usr/bin/env python3
"""캐릭터 상세 명세서 생성 스크립트
데이터 소스: _all_chars.json, Character.xlsx
"""
import json, os, sys
import openpyxl

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(BASE, "data")
DOCS = os.path.join(BASE, "docs")

# ── 1. Load _all_chars.json ──
with open(os.path.join(DATA, "_all_chars.json"), encoding="utf-8") as f:
    all_chars = json.load(f)

# ── 2. Parse Character.xlsx for StarGrade by char_id ──
wb_char = openpyxl.load_workbook(os.path.join(DATA, "Character.xlsx"), data_only=True)
ws_detail = wb_char["Detail<Child>"]

# col 4 = D = char_id (e.g. 'c035'), col 12 = L = StarGrade
star_by_id = {}
for r in range(4, ws_detail.max_row + 1):
    cid = ws_detail.cell(r, 4).value   # D: char_id directly as string 'c035'
    star = ws_detail.cell(r, 12).value  # L: StarGrade 1/2/3
    if cid and star is not None:
        # Normalise: strip whitespace, ensure lowercase 'c###' format
        cid_str = str(cid).strip()
        # If cid came back as a formula result or number, derive from col C
        if not cid_str.startswith("c"):
            col_c = ws_detail.cell(r, 3).value
            if col_c:
                cid_str = "c" + str(int(col_c))[1:4]
            else:
                continue
        star_by_id[cid_str] = int(star)

# ── 3. Translation Maps ──
elem_kr = {"FIRE": "화(火)", "WATER": "수(水)", "FOREST": "목(木)", "LIGHT": "광(光)", "DARK": "암(暗)"}
elem_order = ["FIRE", "WATER", "FOREST", "LIGHT", "DARK"]
elem_icon = {"FIRE": "🔥", "WATER": "💧", "FOREST": "🌿", "LIGHT": "✨", "DARK": "🌑"}
role_kr = {
    "ATTACKER": "공격자(ATK)", "DEFENDER": "방어자(DEF)",
    "HEALER": "힐러(HLR)", "SUPPORTER": "서포터(SUP)", "MAGICIAN": "마법사(MAG)",
}
role_short = {
    "ATTACKER": "ATT", "MAGICIAN": "MAG", "DEFENDER": "DEF",
    "HEALER": "HEA", "SUPPORTER": "SUP",
}
role_icon = {
    "ATTACKER": "⚔️", "DEFENDER": "🛡️", "HEALER": "💚",
    "SUPPORTER": "📣", "MAGICIAN": "🔮",
}
target_kr = {
    "ENEMY_NEAR": "근접 적 1기",
    "ENEMY_RANDOM": "랜덤 적",
    "ENEMY_RANDOM_2": "랜덤 적 2기",
    "ENEMY_RANDOM_3": "랜덤 적 3기",
    "ENEMY_LOWEST_HP": "최저HP 적 1기",
    "ENEMY_NEAR_ROW": "근접 열 적",
    "ENEMY_BACK_ROW": "후열 적",
    "ENEMY_SAME_COL": "같은 열 적",
    "ALL_ENEMY": "적 전체",
    "ALL_ALLY": "아군 전체",
    "ALLY_LOWEST_HP": "최저HP 아군 1기",
    "ALLY_LOWEST_HP_2": "최저HP 아군 2기",
    "ALLY_HIGHEST_ATK": "최고ATK 아군 1기",
    "SELF": "자신",
}

def star_str(cid):
    g = star_by_id.get(cid)
    if g is None:
        return "?"
    return "★" * g

# ── 4. Group and sort chars ──
chars_by_elem = {}
for c in all_chars:
    chars_by_elem.setdefault(c["element"], []).append(c)

role_sort = {"ATTACKER": 0, "MAGICIAN": 1, "DEFENDER": 2, "HEALER": 3, "SUPPORTER": 4}
for elem in chars_by_elem:
    chars_by_elem[elem].sort(key=lambda c: (role_sort.get(c["role"], 9), c["name"]))

# Grade counts for header
grade_counts = {1: 0, 2: 0, 3: 0}
for c in all_chars:
    g = star_by_id.get(c["id"])
    if g in grade_counts:
        grade_counts[g] += 1

# ── 5. Generate Markdown ──
L = []

L.append("# 📋 BattleAgent 캐릭터 상세 명세서")
L.append("")
L.append("**생성일**: 2026-03-16")
L.append("**데이터 소스**: Character.xlsx, Skill.xlsx, _all_chars.json")
L.append(f"**총 캐릭터 수**: {len(all_chars)}명 "
         f"(1성 {grade_counts[1]} / 2성 {grade_counts[2]} / 3성 {grade_counts[3]})")
L.append("**밸런스 룰**: CRI 15% / PEN 0% 통일")
L.append("")

# ── Table of Contents ──
L.append("## 📑 목차")
L.append("")
for elem in elem_order:
    if elem not in chars_by_elem:
        continue
    ic = elem_icon[elem]
    kr = elem_kr[elem]
    chars = chars_by_elem[elem]
    names = ", ".join(c["name"] for c in chars)
    L.append(f"- {ic} **{kr}** ({len(chars)}명): {names}")
L.append("")

# ── Global Summary Table ──
L.append("## 📊 전체 캐릭터 요약")
L.append("")
L.append("| 속성 | 이름 | ID | 역할 | 성급(★) | ATK | DEF | HP | SPD | SP | CRI | PEN |")
L.append("|:---:|------|:---:|:---:|:-------:|----:|----:|-----:|----:|:---:|:---:|:---:|")
for elem in elem_order:
    if elem not in chars_by_elem:
        continue
    for c in chars_by_elem[elem]:
        ss = star_str(c["id"])
        cri_pct = f"{int(c['cri']*100)}%"
        pen_pct = f"{int(c['pen']*100)}%"
        e_ic = elem_icon[c["element"]]
        r_ic = role_icon[c["role"]]
        rs = role_short[c["role"]]
        L.append(
            f"| {e_ic} | **{c['name']}** | {c['id']} | {r_ic} {rs} | "
            f"{ss} | {c['atk']} | {c['def']} | {c['hp']} | {c['spd']} | "
            f"{c['sp']} | {cri_pct} | {pen_pct} |"
        )
L.append("")

# ── Stats Rankings ──
L.append("## 🏆 스탯 랭킹")
L.append("")

sorted_atk = sorted(all_chars, key=lambda c: c["atk"], reverse=True)
sorted_hp  = sorted(all_chars, key=lambda c: c["hp"],  reverse=True)
sorted_def = sorted(all_chars, key=lambda c: c["def"], reverse=True)

L.append("### ATK TOP 10")
L.append("| 순위 | 이름 | 역할 | 속성 | ATK |")
L.append("|:---:|------|:---:|:---:|----:|")
for i, c in enumerate(sorted_atk[:10], 1):
    L.append(f"| {i} | {c['name']} | {role_short[c['role']]} | {elem_icon[c['element']]} | **{c['atk']}** |")
L.append("")

L.append("### HP TOP 10")
L.append("| 순위 | 이름 | 역할 | 속성 | HP |")
L.append("|:---:|------|:---:|:---:|-----:|")
for i, c in enumerate(sorted_hp[:10], 1):
    L.append(f"| {i} | {c['name']} | {role_short[c['role']]} | {elem_icon[c['element']]} | **{c['hp']}** |")
L.append("")

L.append("### DEF TOP 10")
L.append("| 순위 | 이름 | 역할 | 속성 | DEF |")
L.append("|:---:|------|:---:|:---:|----:|")
for i, c in enumerate(sorted_def[:10], 1):
    L.append(f"| {i} | {c['name']} | {role_short[c['role']]} | {elem_icon[c['element']]} | **{c['def']}** |")
L.append("")

L.append("### SPD 분포")
L.append("| SPD | 역할군 | 캐릭터 |")
L.append("|:---:|:------:|--------|")
spd_groups = {}
for c in all_chars:
    spd_groups.setdefault(c["spd"], []).append(c)
for spd in sorted(spd_groups.keys(), reverse=True):
    chars_in = spd_groups[spd]
    roles = sorted(set(role_short[c["role"]] for c in chars_in))
    names = ", ".join(c["name"] for c in chars_in)
    L.append(f"| **{spd}** | {'/'.join(roles)} | {names} |")
L.append("")

# ── Per-element Per-character Detail ──
for elem in elem_order:
    if elem not in chars_by_elem:
        continue
    e_ic = elem_icon[elem]
    e_kr = elem_kr[elem]
    L.append("---")
    L.append(f"## {e_ic} {e_kr} 속성 캐릭터")
    L.append("")

    for c in chars_by_elem[elem]:
        ss = star_str(c["id"])
        r_ic = role_icon[c["role"]]
        r_kr = role_kr[c["role"]]

        L.append(f"### {r_ic} {c['name']} ({c['id']})")
        L.append("")
        L.append(f"> **{e_ic} {e_kr}** | **{r_kr}** | **등급: {ss}** | **SP 코스트: {c['sp']}**")
        L.append("")

        # Stats table
        L.append("#### 기본 스탯")
        L.append("")
        L.append("| ATK | DEF | HP | SPD | CRI | CRI_DMG | PEN |")
        L.append("|----:|----:|-----:|----:|:---:|:-------:|:---:|")
        cri_pct = f"{int(c['cri']*100)}%"
        pen_pct = f"{int(c['pen']*100)}%"
        L.append(f"| {c['atk']} | {c['def']} | {c['hp']} | {c['spd']} | {cri_pct} | 150% | {pen_pct} |")
        L.append("")

        # SP system
        L.append(f"**SP 시스템**: MaxSP={c['sp']} / UseSP={c['sp']} / TurnSP=+1")
        L.append("")

        # Skills
        L.append("#### 스킬 상세")
        L.append("")

        has_skills = c.get("normal") and c.get("active") and c.get("ultimate")

        if has_skills:
            n = c["normal"]
            t_kr = target_kr.get(n["target"], n["target"])
            L.append(f"**🟢 노멀 — {n['name']}**")
            L.append(f"- 타겟: {t_kr}")
            L.append(f"- 효과: {n['effects']}")
            L.append("")

            a = c["active"]
            t_kr = target_kr.get(a["target"], a["target"])
            L.append(f"**🔵 액티브 — {a['name']}**")
            L.append(f"- 타겟: {t_kr}")
            L.append(f"- 효과: {a['effects']}")
            L.append("")

            u = c["ultimate"]
            t_kr = target_kr.get(u["target"], u["target"])
            L.append(f"**🟣 얼티밋 — {u['name']}** (SP {c['sp']})")
            L.append(f"- 타겟: {t_kr}")
            L.append(f"- 효과: {u['effects']}")
            L.append("")
        else:
            L.append("**스킬 데이터 미등록**")
            L.append("")

# ── Role Distribution ──
L.append("---")
L.append("## 📈 역할별 분포 분석")
L.append("")

role_counts = {}
for c in all_chars:
    role_counts.setdefault(c["role"], {"count": 0, "chars": []})
    role_counts[c["role"]]["count"] += 1
    role_counts[c["role"]]["chars"].append(c["name"])

L.append("| 역할 | 수 | 캐릭터 |")
L.append("|:---:|:---:|--------|")
for role in ["ATTACKER", "MAGICIAN", "DEFENDER", "HEALER", "SUPPORTER"]:
    if role in role_counts:
        info = role_counts[role]
        L.append(f"| {role_icon[role]} {role_kr[role]} | {info['count']} | {', '.join(info['chars'])} |")
L.append("")

# ── Element x Role Distribution ──
L.append("## 🌈 속성별 역할 분포")
L.append("")
L.append("| 속성 | 총 수 | ⚔️ATT | 🔮MAG | 🛡️DEF | 💚HEA | 📣SUP |")
L.append("|:---:|:---:|:---:|:---:|:---:|:---:|:---:|")
for elem in elem_order:
    if elem not in chars_by_elem:
        continue
    chars = chars_by_elem[elem]
    rd = {}
    for ch in chars:
        rd[ch["role"]] = rd.get(ch["role"], 0) + 1
    L.append(
        f"| {elem_icon[elem]} {elem_kr[elem]} | {len(chars)} | "
        f"{rd.get('ATTACKER',0)} | {rd.get('MAGICIAN',0)} | "
        f"{rd.get('DEFENDER',0)} | {rd.get('HEALER',0)} | "
        f"{rd.get('SUPPORTER',0)} |"
    )
L.append("")

# ── Avg Stats by Role ──
L.append("## 📉 역할별 평균 스탯 비교")
L.append("")
L.append("| 역할 | 수 | 평균ATK | 평균DEF | 평균HP | 평균SPD |")
L.append("|:---:|:---:|------:|------:|------:|------:|")
for role in ["ATTACKER", "MAGICIAN", "DEFENDER", "HEALER", "SUPPORTER"]:
    chars_r = [c for c in all_chars if c["role"] == role]
    if not chars_r:
        continue
    n = len(chars_r)
    avg_atk = sum(c["atk"] for c in chars_r) / n
    avg_def = sum(c["def"] for c in chars_r) / n
    avg_hp  = sum(c["hp"]  for c in chars_r) / n
    avg_spd = sum(c["spd"] for c in chars_r) / n
    L.append(f"| {role_icon[role]} {role_kr[role]} | {n} | {avg_atk:.0f} | {avg_def:.0f} | {avg_hp:.0f} | {avg_spd:.0f} |")
L.append("")

# ── Write output ──
output = "\n".join(L)
os.makedirs(DOCS, exist_ok=True)
out_path = os.path.join(DOCS, "캐릭터_상세_명세서.md")
with open(out_path, "w", encoding="utf-8") as f:
    f.write(output)

print(f"완료: {out_path}")
print(f"총 {len(all_chars)}명 캐릭터, {len(L)}줄")
print(f"StarGrade 매핑: {len(star_by_id)}개 ID")
print(f"1성 {grade_counts[1]} / 2성 {grade_counts[2]} / 3성 {grade_counts[3]}")
