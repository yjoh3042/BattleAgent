"""스탯 규약에 맞춰 모든 캐릭터 스탯 정규화.

Usage:
  py -3 -X utf8 scripts/normalize_stats.py          # 분석만
  py -3 -X utf8 scripts/normalize_stats.py --fix     # 분석 + 엑셀 수정
"""
import math
import sys
from openpyxl import load_workbook

STANDARD = {
    'Attacker': {
        'Hp':  {1: 2400, 2: 3200, 3: 4000, 3.5: 4600},
        'Atk': {1: 240,  2: 320,  3: 400,  3.5: 460},
        'Def': {1: 120,  2: 160,  3: 200,  3.5: 230},
    },
    'Defender': {
        'Hp':  {1: 3600, 2: 4800, 3: 6000, 3.5: 6900},
        'Atk': {1: 150,  2: 200,  3: 250,  3.5: 287},
        'Def': {1: 180,  2: 240,  3: 300,  3.5: 345},
    },
    'Magician': {
        'Hp':  {1: 3000, 2: 4000, 3: 5000, 3.5: 5750},
        'Atk': {1: 180,  2: 240,  3: 300,  3.5: 345},
        'Def': {1: 120,  2: 160,  3: 200,  3.5: 230},
    },
    'Supporter': {
        'Hp':  {1: 3000, 2: 4000, 3: 5000, 3.5: 5750},
        'Atk': {1: 150,  2: 200,  3: 250,  3.5: 287},
        'Def': {1: 120,  2: 160,  3: 200,  3.5: 230},
    },
    'Healer': {
        'Hp':  {1: 2400, 2: 3200, 3: 4000, 3.5: 4600},
        'Atk': {1: 120,  2: 160,  3: 200,  3.5: 230},
        'Def': {1: 120,  2: 160,  3: 200,  3.5: 230},
    },
}

STAT_COLS = list(range(21, 39, 2))  # StatParam columns (U,W,Y,AA,AC,AE,AG,AI,AK)
FIX_MODE = '--fix' in sys.argv
EXCEL_FILES = ['data/Character.xlsx', 'data/Character_updated.xlsx']


def analyze_and_fix(path):
    wb = load_workbook(path)
    ws = wb['Detail<Child>']

    print(f"=== {path} ===")
    print(f"{'이름':>8} | {'등급':>4} | {'직업':>10} | {'HP현재':>6} | {'HP규약':>6} | {'ATK현':>5} | {'ATK규':>5} | {'DEF현':>5} | {'DEF규':>5} | 상태")
    print("-" * 105)

    total = 0
    changed_count = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        id_val = row[0].value
        if not id_val or not isinstance(id_val, (int, float)):
            continue
        name = row[1].value
        grade_raw = row[11].value
        role = row[13].value
        if not role or not grade_raw:
            continue
        grade = float(grade_raw)
        if grade not in (1, 2, 3, 3.5):
            continue
        if role not in STANDARD:
            continue

        total += 1
        cur = {}
        cell_map = {}
        for ci in STAT_COLS:
            pc = row[ci - 1]
            vc = row[ci]
            if pc.value in ('Hp', 'Atk', 'Def'):
                v = vc.value
                if v is not None and not isinstance(v, str):
                    cur[pc.value] = int(v)
                    cell_map[pc.value] = vc

        std = STANDARD[role]
        hp_s = std['Hp'][grade]
        atk_s = std['Atk'][grade]
        def_s = std['Def'][grade]
        hp_c = cur.get('Hp', 0)
        atk_c = cur.get('Atk', 0)
        def_c = cur.get('Def', 0)

        need_change = (hp_c != hp_s) or (atk_c != atk_s) or (def_c != def_s)
        mark = "*** 변경필요" if need_change else "OK"
        if need_change:
            changed_count += 1

        print(f"{name:>8} | {grade:>4} | {role:>10} | {hp_c:>6} | {hp_s:>6} | {atk_c:>5} | {atk_s:>5} | {def_c:>5} | {def_s:>5} | {mark}")

        if FIX_MODE and need_change:
            if 'Hp' in cell_map:
                cell_map['Hp'].value = hp_s
            if 'Atk' in cell_map:
                cell_map['Atk'].value = atk_s
            if 'Def' in cell_map:
                cell_map['Def'].value = def_s

    print(f"\n총 {total}명 중 {changed_count}명 변경 필요")

    if FIX_MODE and changed_count > 0:
        wb.save(path)
        print(f">> {path} 저장 완료 ({changed_count}명 수정)")
    print()


if __name__ == '__main__':
    for f in EXCEL_FILES:
        analyze_and_fix(f)
