"""c600 루미나 / c601 에레보스 스킬 상세 기획서 xlsx 생성"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
hdr_fill = PatternFill('solid', fgColor='4472C4')
hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
sub_fill = PatternFill('solid', fgColor='D6E4F0')
sub_font = Font(name='Arial', bold=True, size=10)
data_font = Font(name='Arial', size=10)
thin_border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
center = Alignment(horizontal='center', vertical='center', wrap_text=True)

def sh(ws, row, mc):
    for c in range(1, mc+1):
        cell = ws.cell(row=row, column=c)
        cell.fill, cell.font, cell.alignment, cell.border = hdr_fill, hdr_font, center, thin_border

def sd(ws, row, mc):
    for c in range(1, mc+1):
        cell = ws.cell(row=row, column=c)
        cell.font, cell.alignment, cell.border = data_font, center, thin_border

# ===== Sheet 1: 기획 개요 =====
ws1 = wb.active
ws1.title = '기획 개요'
ws1.column_dimensions['A'].width = 18
ws1.column_dimensions['B'].width = 40
ws1.column_dimensions['C'].width = 18
ws1.column_dimensions['D'].width = 40
ws1['A1'] = '신규 캐릭터 스킬 상세 기획서'
ws1['A1'].font = Font(name='Arial', bold=True, size=16)
ws1.merge_cells('A1:D1')
overview = [
    ['', 'c600 루미나', '', 'c601 에레보스'],
    ['ID', '10000600', 'ID', '10000601'],
    ['속성', '광 (Light)', '속성', '목 (Forest)'],
    ['역할', '서포터 (Supporter)', '역할', '공격자 (Attacker)'],
    ['성급', '5성', '성급', '5성'],
    ['ATK', '230', 'ATK', '530'],
    ['DEF', '160', 'DEF', '135'],
    ['HP', '6800', 'HP', '5500'],
    ['SPD', '120', 'SPD', '80'],
    ['CRI', '5% (기본)', 'CRI', '35%'],
    ['PEN', '0% (기본)', 'PEN', '20%'],
    ['MaxSP / UseSP', '4 / 3', 'MaxSP / UseSP', '6 / 5'],
    ['', '', '', ''],
    ['설계 의도', '시트리+엘리시온+브라우니 통합', '설계 의도', '이브+쿠바바+아르테미스 통합'],
    ['OP 포인트1', 'Active 전체ATK+30%,SPD+20,정화', 'OP 포인트1', 'ATK 530 + CRI 35% + PEN 20%'],
    ['OP 포인트2', 'Ult 힐35%+ATK35%+CriDmg50%+SP1', 'OP 포인트2', 'Normal 2연타 실질 5.0x'],
    ['OP 포인트3', 'SP3으로 저렴한 궁극기', 'OP 포인트3', 'Active HP<50% 처형 총8.0x'],
]
for i, row in enumerate(overview, 3):
    for j, val in enumerate(row):
        cell = ws1.cell(row=i, column=j+1, value=val)
        cell.font = data_font
        cell.border = thin_border
        if j in (0, 2):
            cell.font = sub_font
            cell.fill = sub_fill
for c in range(1, 5):
    ws1.cell(row=3, column=c).fill = hdr_fill
    ws1.cell(row=3, column=c).font = hdr_font

# ===== Sheet 2: Skill<Child> =====
ws2 = wb.create_sheet('Skill(Child)')
skill_cols = [
    'id','#Character','#SkillName','#SkillDesc','Name','SkillAtlas','SkillIcon',
    'Desc','TargetAtlas','TargetIcon','TargetAreaAtlas','TargetAreaIcon',
    'Element','Role','Type','Cooltime','SkillCamType','CastingType','CastingOffset',
    'ActionTriggerType1','RemoveTriggerType1','RemoveTriggerParam1','Target1','BuffValue1',
    'ActionTriggerType2','RemoveTriggerType2','RemoveTriggerParam2','Target2','BuffValue2',
    'ActionTriggerType3','RemoveTriggerType3','RemoveTriggerParam3','Target3','BuffValue3',
    'ActionTriggerType4','RemoveTriggerType4','RemoveTriggerParam4','Target4','BuffValue4',
    '#MainTarget','#Main','#MainValue',
    '#SubTarget1','#Sub1','#SubValue1',
    '#SubTarget2','#Sub2','#SubValue2',
    '#SubTarget3','#Sub3','#SubValue3',
]
for c, name in enumerate(skill_cols, 1):
    ws2.cell(row=1, column=c, value=name)
    ws2.column_dimensions[get_column_letter(c)].width = max(len(name)+2, 12)
sh(ws2, 1, len(skill_cols))

E = ''
c600s = [
    [1544601,'c600','루미나_노멀','성광의 채찍','c600_normal','SkillAtlas_c600','skill_c600_n',
     '신성한 빛으로 적 1명을 공격','TargetAtlas','target_single','AreaAtlas','area_single',
     4,4,'Normal',0,'Normal','Move',E,
     'SignalHit',E,E,2000011,15446011, E,E,E,E,E, E,E,E,E,E, E,E,E,E,E,
     '적1','공격','120%', E,E,E, E,E,E, E,E,E],
    [2544601,'c600','루미나_액티브1','빛의 축복','c600_active1','SkillAtlas_c600','skill_c600_a',
     '아군 전체 ATK+30%, SPD+20, 디버프 제거','TargetAtlas','target_all_ally','AreaAtlas','area_all',
     4,4,'Active',3,'Buff','Static',E,
     'Immediately',E,E,1000000,25446011,
     'Immediately',E,E,1000000,25446012,
     'Immediately',E,E,1000000,25446013,
     E,E,E,E,E,
     '아군전체','ATK+30%','3턴', '아군전체','SPD+20','3턴', '아군전체','디버프제거','즉시'],
    [2544602,'c600','루미나_액티브2','빛의 축복+','c600_active2','SkillAtlas_c600','skill_c600_a',
     '아군 전체 ATK+35%, SPD+25, 디버프 제거','TargetAtlas','target_all_ally','AreaAtlas','area_all',
     4,4,'Active',3,'Buff','Static',E,
     'Immediately',E,E,1000000,25446021,
     'Immediately',E,E,1000000,25446022,
     'Immediately',E,E,1000000,25446023,
     E,E,E,E,E,
     '아군전체','ATK+35%','3턴', '아군전체','SPD+25','3턴', '아군전체','디버프제거','즉시'],
    [3544601,'c600','루미나_얼티1','성역의 은총','c600_ult1','SkillAtlas_c600','skill_c600_u',
     '아군 전체 HP35%회복+ATK+35%+CriDmg+50%+SP+1','TargetAtlas','target_all_ally','AreaAtlas','area_all',
     4,4,'Ultimate',0,'Buff','Static',E,
     'Immediately',E,E,1000000,35446011,
     'Immediately',E,E,1000000,35446012,
     'Immediately',E,E,1000000,35446013,
     'Immediately',E,E,1000000,35446014,
     '아군전체','회복35%','즉시', '아군전체','ATK+35%','3턴', '아군전체','CriDmg+50%','3턴', '아군전체','SP+1','즉시'],
    [3544602,'c600','루미나_얼티2','성역의 은총+','c600_ult2','SkillAtlas_c600','skill_c600_u',
     '아군 전체 HP40%회복+ATK+40%+CriDmg+55%+SP+1','TargetAtlas','target_all_ally','AreaAtlas','area_all',
     4,4,'Ultimate',0,'Buff','Static',E,
     'Immediately',E,E,1000000,35446021,
     'Immediately',E,E,1000000,35446022,
     'Immediately',E,E,1000000,35446023,
     'Immediately',E,E,1000000,35446024,
     '아군전체','회복40%','즉시', '아군전체','ATK+40%','3턴', '아군전체','CriDmg+55%','3턴', '아군전체','SP+1','즉시'],
    [3544603,'c600','루미나_얼티3','성역의 은총++','c600_ult3','SkillAtlas_c600','skill_c600_u',
     '아군 전체 HP45%회복+ATK+45%+CriDmg+60%+SP+2','TargetAtlas','target_all_ally','AreaAtlas','area_all',
     4,4,'Ultimate',0,'Buff','Static',E,
     'Immediately',E,E,1000000,35446031,
     'Immediately',E,E,1000000,35446032,
     'Immediately',E,E,1000000,35446033,
     'Immediately',E,E,1000000,35446034,
     '아군전체','회복45%','즉시', '아군전체','ATK+45%','3턴', '아군전체','CriDmg+60%','3턴', '아군전체','SP+2','즉시'],
    [4544601,'c600','루미나_패시브1','빛의 가호','c600_passive1','SkillAtlas_c600','skill_c600_p1',
     '전투 시작 시 아군 전체 DEF+15%(영구)','TargetAtlas','target_all_ally','AreaAtlas','area_all',
     4,4,'Passive',0,'None','Static',E,
     'BattleStart',E,E,1000000,45446011, E,E,E,E,E, E,E,E,E,E, E,E,E,E,E,
     '아군전체','DEF+15%','영구', E,E,E, E,E,E, E,E,E],
    [4544602,'c600','루미나_패시브2','성스러운 빛','c600_passive2','SkillAtlas_c600','skill_c600_p2',
     '아군 피격시 5% 확률로 HP 10% 회복','TargetAtlas','target_self','AreaAtlas','area_single',
     4,4,'Passive',0,'None','Static',E,
     'OnAllyDamaged',E,E,1200001,45446021, E,E,E,E,E, E,E,E,E,E, E,E,E,E,E,
     '피격아군','회복10%','5%확률', E,E,E, E,E,E, E,E,E],
]

c601s = [
    [1531601,'c601','에레보스_노멀','심연의 쌍격','c601_normal','SkillAtlas_c601','skill_c601_n',
     '암흑의 힘으로 적 1명을 2회 공격(각 250%)','TargetAtlas','target_single','AreaAtlas','area_single',
     3,1,'Normal',0,'Normal','Move',E,
     'SignalHit',E,E,2000011,15316011,
     'SignalHit',E,E,2000011,15316012,
     E,E,E,E,E, E,E,E,E,E,
     '적1','공격250%','1타', '적1','공격250%','2타', E,E,E, E,E,E],
    [2531601,'c601','에레보스_액티브1','종결자의 일격','c601_active1','SkillAtlas_c601','skill_c601_a',
     '최저HP적 500%공격+HP<50%시 추가300%+DEF-25%','TargetAtlas','target_lowest','AreaAtlas','area_single',
     3,1,'Active',4,'Attack','Move',E,
     'SignalHit',E,E,2000012,25316011,
     'SignalHit',10201,50000,2000012,25316012,
     'AttackEnd',E,E,2000012,25316013,
     E,E,E,E,E,
     '적최저HP','공격500%','즉시', '적최저HP','추가300%','HP<50%', '적최저HP','DEF-25%','2턴', E,E,E],
    [2531602,'c601','에레보스_액티브2','종결자의 일격+','c601_active2','SkillAtlas_c601','skill_c601_a',
     '최저HP적 600%공격+HP<50%시 추가350%+DEF-30%','TargetAtlas','target_lowest','AreaAtlas','area_single',
     3,1,'Active',4,'Attack','Move',E,
     'SignalHit',E,E,2000012,25316021,
     'SignalHit',10201,50000,2000012,25316022,
     'AttackEnd',E,E,2000012,25316023,
     E,E,E,E,E,
     '적최저HP','공격600%','즉시', '적최저HP','추가350%','HP<50%', '적최저HP','DEF-30%','2턴', E,E,E],
    [3531601,'c601','에레보스_얼티1','심판의 심연','c601_ult1','SkillAtlas_c601','skill_c601_u',
     '적 전체 350%공격+DEF-30%(2턴)+출혈DoT 20%(3턴)','TargetAtlas','target_all_enemy','AreaAtlas','area_all',
     3,1,'Ultimate',0,'Attack','NonTargetMove',E,
     'SignalHit',E,E,920001,35316011,
     'AttackEnd',E,E,920001,35316012,
     'AttackEnd',E,E,920001,35316013,
     E,E,E,E,E,
     '적전체','공격350%','즉시', '적전체','DEF-30%','2턴', '적전체','출혈20%','3턴', E,E,E],
    [3531602,'c601','에레보스_얼티2','심판의 심연+','c601_ult2','SkillAtlas_c601','skill_c601_u',
     '적 전체 400%공격+DEF-35%(2턴)+출혈DoT 25%(3턴)','TargetAtlas','target_all_enemy','AreaAtlas','area_all',
     3,1,'Ultimate',0,'Attack','NonTargetMove',E,
     'SignalHit',E,E,920001,35316021,
     'AttackEnd',E,E,920001,35316022,
     'AttackEnd',E,E,920001,35316023,
     E,E,E,E,E,
     '적전체','공격400%','즉시', '적전체','DEF-35%','2턴', '적전체','출혈25%','3턴', E,E,E],
    [3531603,'c601','에레보스_얼티3','심판의 심연++','c601_ult3','SkillAtlas_c601','skill_c601_u',
     '적 전체 450%공격+DEF-40%(2턴)+출혈DoT 30%(3턴)','TargetAtlas','target_all_enemy','AreaAtlas','area_all',
     3,1,'Ultimate',0,'Attack','NonTargetMove',E,
     'SignalHit',E,E,920001,35316031,
     'AttackEnd',E,E,920001,35316032,
     'AttackEnd',E,E,920001,35316033,
     E,E,E,E,E,
     '적전체','공격450%','즉시', '적전체','DEF-40%','2턴', '적전체','출혈30%','3턴', E,E,E],
    [4531601,'c601','에레보스_패시브1','심연의 힘','c601_passive1','SkillAtlas_c601','skill_c601_p1',
     '전투 시작 시 자신 ATK+20%+PEN+10%(영구)','TargetAtlas','target_self','AreaAtlas','area_self',
     3,1,'Passive',0,'None','Static',E,
     'BattleStart',E,E,1100000,45316011,
     'BattleStart',E,E,1100000,45316012,
     E,E,E,E,E, E,E,E,E,E,
     '자신','ATK+20%','영구', '자신','PEN+10%','영구', E,E,E, E,E,E],
    [4531602,'c601','에레보스_패시브2','처형자의 본능','c601_passive2','SkillAtlas_c601','skill_c601_p2',
     '적 처치 시 ATK+15%(2턴)','TargetAtlas','target_self','AreaAtlas','area_self',
     3,1,'Passive',0,'None','Static',E,
     'OnKill',E,E,1100000,45316021, E,E,E,E,E, E,E,E,E,E, E,E,E,E,E,
     '자신','ATK+15%','2턴(처치시)', E,E,E, E,E,E, E,E,E],
]

for i, row in enumerate(c600s + c601s, 2):
    for j, val in enumerate(row):
        ws2.cell(row=i, column=j+1, value=val)
    sd(ws2, i, len(skill_cols))
print('Sheet 2 done')

# ===== Sheet 3: BuffValue<Child> =====
ws3 = wb.create_sheet('BuffValue(Child)')
bv_cols = [
    'Id','#Character','#Type','#Skilldesc','Buff','#BuffDesc',
    'BuffTrigger','RequireTag1','RequireTag2','BuffRate',
    'V1_L1','V2_L1','V3_L1','V1_L2','V2_L2','V3_L2','V1_L3','V2_L3','V3_L3',
    'BuffTurn',
    'LvUp1_L1','LvUp2_L1','LvUp3_L1','LvUp1_L2','LvUp2_L2','LvUp3_L2','LvUp1_L3','LvUp2_L3','LvUp3_L3',
]
for c, name in enumerate(bv_cols, 1):
    ws3.cell(row=1, column=c, value=name)
    ws3.column_dimensions[get_column_letter(c)].width = max(len(name)+2, 11)
sh(ws3, 1, len(bv_cols))

Z = [0]*9  # 9 zeros for LvUp columns
c600bv = [
    [15446011,'c600','Normal','성광의채찍_대미지',1000001,'공용대미지',E,E,E,100000, 120000,0,0, 0,0,0, 0,0,0, 0]+Z,
    [25446011,'c600','Active1','빛의축복_ATK',2000101,'ATK증가(%)',E,E,E,100000, 30000,1001,0, 0,0,0, 0,0,0, 3]+Z,
    [25446012,'c600','Active1','빛의축복_SPD',2001001,'SPD증가',E,E,E,100000, 20000,1010,0, 0,0,0, 0,0,0, 3]+Z,
    [25446013,'c600','Active1','빛의축복_정화',4000101,'디버프제거',E,E,E,100000, 1,0,0, 0,0,0, 0,0,0, 0]+Z,
    [25446021,'c600','Active2','빛의축복+_ATK',2000101,'ATK증가(%)',E,E,E,100000, 35000,1001,0, 0,0,0, 0,0,0, 3]+Z,
    [25446022,'c600','Active2','빛의축복+_SPD',2001001,'SPD증가',E,E,E,100000, 25000,1010,0, 0,0,0, 0,0,0, 3]+Z,
    [25446023,'c600','Active2','빛의축복+_정화',4000101,'디버프제거',E,E,E,100000, 1,0,0, 0,0,0, 0,0,0, 0]+Z,
    [35446011,'c600','Ult1','성역의은총_힐',9900003,'루미나회복',E,E,E,100000, 35000,0,0, 0,0,0, 0,0,0, 0]+Z,
    [35446012,'c600','Ult1','성역의은총_ATK',2000101,'ATK증가(%)',E,E,E,100000, 35000,1001,0, 0,0,0, 0,0,0, 3]+Z,
    [35446013,'c600','Ult1','성역의은총_CriDmg',2000501,'CriDmg증가',E,E,E,100000, 50000,1005,0, 0,0,0, 0,0,0, 3]+Z,
    [35446014,'c600','Ult1','성역의은총_SP',2003001,'SP회복',E,E,E,100000, 1,0,0, 0,0,0, 0,0,0, 0]+Z,
    [35446021,'c600','Ult2','성역의은총+_힐',9900003,'루미나회복',E,E,E,100000, 40000,0,0, 0,0,0, 0,0,0, 0]+Z,
    [35446022,'c600','Ult2','성역의은총+_ATK',2000101,'ATK증가(%)',E,E,E,100000, 40000,1001,0, 0,0,0, 0,0,0, 3]+Z,
    [35446023,'c600','Ult2','성역의은총+_CriDmg',2000501,'CriDmg증가',E,E,E,100000, 55000,1005,0, 0,0,0, 0,0,0, 3]+Z,
    [35446024,'c600','Ult2','성역의은총+_SP',2003001,'SP회복',E,E,E,100000, 1,0,0, 0,0,0, 0,0,0, 0]+Z,
    [35446031,'c600','Ult3','성역의은총++_힐',9900003,'루미나회복',E,E,E,100000, 45000,0,0, 0,0,0, 0,0,0, 0]+Z,
    [35446032,'c600','Ult3','성역의은총++_ATK',2000101,'ATK증가(%)',E,E,E,100000, 45000,1001,0, 0,0,0, 0,0,0, 3]+Z,
    [35446033,'c600','Ult3','성역의은총++_CriDmg',2000501,'CriDmg증가',E,E,E,100000, 60000,1005,0, 0,0,0, 0,0,0, 3]+Z,
    [35446034,'c600','Ult3','성역의은총++_SP',2003001,'SP회복',E,E,E,100000, 2,0,0, 0,0,0, 0,0,0, 0]+Z,
    [45446011,'c600','Passive1','빛의가호_DEF',2000201,'DEF증가(%)',E,E,E,100000, 15000,1003,0, 0,0,0, 0,0,0, -1]+Z,
    [45446021,'c600','Passive2','성스러운빛_힐',9900003,'루미나회복',E,E,E,5000, 10000,0,0, 0,0,0, 0,0,0, 0]+Z,
]

c601bv = [
    [15316011,'c601','Normal','심연의쌍격_1타',1000001,'공용대미지',E,E,E,100000, 250000,0,0, 0,0,0, 0,0,0, 0]+Z,
    [15316012,'c601','Normal','심연의쌍격_2타',1000001,'공용대미지',E,E,E,100000, 250000,0,0, 0,0,0, 0,0,0, 0]+Z,
    [25316011,'c601','Active1','종결자의일격_메인',1000002,'액티브대미지',E,E,E,100000, 500000,0,0, 0,0,0, 0,0,0, 0]+Z,
    [25316012,'c601','Active1','종결자의일격_처형',1000015,'조건부대미지',10201,50000,E,100000, 300000,0,0, 0,0,0, 0,0,0, 0]+Z,
    [25316013,'c601','Active1','종결자의일격_DEF감소',3000201,'DEF감소(%)',E,E,E,100000, -25000,1003,0, 0,0,0, 0,0,0, 2]+Z,
    [25316021,'c601','Active2','종결자의일격+_메인',1000002,'액티브대미지',E,E,E,100000, 600000,0,0, 0,0,0, 0,0,0, 0]+Z,
    [25316022,'c601','Active2','종결자의일격+_처형',1000015,'조건부대미지',10201,50000,E,100000, 350000,0,0, 0,0,0, 0,0,0, 0]+Z,
    [25316023,'c601','Active2','종결자의일격+_DEF감소',3000201,'DEF감소(%)',E,E,E,100000, -30000,1003,0, 0,0,0, 0,0,0, 2]+Z,
    [35316011,'c601','Ult1','심판의심연_대미지',1000003,'얼티밋대미지',E,E,E,100000, 350000,0,0, 0,0,0, 0,0,0, 0]+Z,
    [35316012,'c601','Ult1','심판의심연_DEF감소',3000201,'DEF감소(%)',E,E,E,100000, -30000,1003,0, 0,0,0, 0,0,0, 2]+Z,
    [35316013,'c601','Ult1','심판의심연_출혈',3002306,'에레보스출혈',E,E,E,100000, 20000,0,0, 0,0,0, 0,0,0, 3]+Z,
    [35316021,'c601','Ult2','심판의심연+_대미지',1000003,'얼티밋대미지',E,E,E,100000, 400000,0,0, 0,0,0, 0,0,0, 0]+Z,
    [35316022,'c601','Ult2','심판의심연+_DEF감소',3000201,'DEF감소(%)',E,E,E,100000, -35000,1003,0, 0,0,0, 0,0,0, 2]+Z,
    [35316023,'c601','Ult2','심판의심연+_출혈',3002306,'에레보스출혈',E,E,E,100000, 25000,0,0, 0,0,0, 0,0,0, 3]+Z,
    [35316031,'c601','Ult3','심판의심연++_대미지',1000003,'얼티밋대미지',E,E,E,100000, 450000,0,0, 0,0,0, 0,0,0, 0]+Z,
    [35316032,'c601','Ult3','심판의심연++_DEF감소',3000201,'DEF감소(%)',E,E,E,100000, -40000,1003,0, 0,0,0, 0,0,0, 2]+Z,
    [35316033,'c601','Ult3','심판의심연++_출혈',3002306,'에레보스출혈',E,E,E,100000, 30000,0,0, 0,0,0, 0,0,0, 3]+Z,
    [45316011,'c601','Passive1','심연의힘_ATK',2000101,'ATK증가(%)',E,E,E,100000, 20000,1001,0, 0,0,0, 0,0,0, -1]+Z,
    [45316012,'c601','Passive1','심연의힘_PEN',2009001,'관통력증가',E,E,E,100000, 10000,0,0, 0,0,0, 0,0,0, -1]+Z,
    [45316021,'c601','Passive2','처형자의본능_ATK',2000101,'ATK증가(%)',E,E,E,100000, 15000,1001,0, 0,0,0, 0,0,0, 2]+Z,
]

for i, row in enumerate(c600bv + c601bv, 2):
    for j, val in enumerate(row):
        ws3.cell(row=i, column=j+1, value=val)
    sd(ws3, i, len(bv_cols))
print('Sheet 3 done')

# ===== Sheet 4: Character =====
ws4 = wb.create_sheet('Character')
char_cols = [
    'Id','#이름','Skin','StarGrade','Element','Role',
    'MaxSp','UseSp','TurnSp',
    'StatParam1','StatValue1','StatParam2','StatValue2',
    'StatParam3','StatValue3','StatParam4','StatValue4',
    'StatParam5','StatValue5','StatParam6','StatValue6',
    'StatParam7','StatValue7','StatParam8','StatValue8',
    'StatParam9','StatValue9',
    'SkillId1','SkillId2','SkillId3','SkillId4',
]
for c, name in enumerate(char_cols, 1):
    ws4.cell(row=1, column=c, value=name)
    ws4.column_dimensions[get_column_letter(c)].width = max(len(name)+2, 12)
sh(ws4, 1, len(char_cols))

chars = [
    [10000600,'루미나','skin_c600',5,4,4, 4,3,1,
     'Spd',120,'Hp',6800,'Atk',230,'Def',160,
     'Penetration',0,'Hit',0,'Dodge',0,'CriRatio',5000,'CriDmgRatio',50000,
     1544601,2544601,3544601,4544601],
    [10000601,'에레보스','skin_c601',5,3,1, 6,5,1,
     'Spd',80,'Hp',5500,'Atk',530,'Def',135,
     'Penetration',20000,'Hit',0,'Dodge',0,'CriRatio',35000,'CriDmgRatio',50000,
     1531601,2531601,3531601,4531601],
]
for i, row in enumerate(chars, 2):
    for j, val in enumerate(row):
        ws4.cell(row=i, column=j+1, value=val)
    sd(ws4, i, len(char_cols))
print('Sheet 4 done')

# ===== Sheet 5: Buff_New =====
ws5 = wb.create_sheet('Buff_New')
buff_cols = ['BuffId','이름','설명','카테고리','대상','비고']
for c, name in enumerate(buff_cols, 1):
    ws5.cell(row=1, column=c, value=name)
    ws5.column_dimensions[get_column_letter(c)].width = 20
sh(ws5, 1, len(buff_cols))

buffs = [
    [9900003,'루미나 전용 회복','HP% 비례 회복','회복','아군','기존 9900001/2 패턴'],
    [4000101,'디버프 제거','대상의 모든 디버프 제거','정화','아군','신규 카테고리'],
    [3002306,'에레보스 전용 출혈','ATK% 비례 DoT (매턴)','DoT','적','기존 3002301~5 패턴'],
    [2009001,'관통력 증가','관통력 수치 증가','버프','자신','신규 스탯 버프'],
    [1000015,'조건부 대미지','HP조건 충족시 추가 대미지','대미지','적','이브 처형 패턴 확장'],
]
for i, row in enumerate(buffs, 2):
    for j, val in enumerate(row):
        ws5.cell(row=i, column=j+1, value=val)
    sd(ws5, i, len(buff_cols))
print('Sheet 5 done')

# ===== Sheet 6: 스킬 상세 설명 =====
ws6 = wb.create_sheet('스킬 상세 설명')
dc = ['캐릭터','스킬타입','스킬명','Lv','대상','효과 요약','수치 상세','특이사항']
widths6 = [10,10,16,5,14,35,40,30]
for c, name in enumerate(dc, 1):
    ws6.cell(row=1, column=c, value=name)
    ws6.column_dimensions[get_column_letter(c)].width = widths6[c-1]
sh(ws6, 1, len(dc))

descs = [
    ['루미나','Normal','성광의 채찍','—','적1','빛 속성 단일 공격','ATK×120%','기본 서포터 노멀'],
    ['루미나','Active','빛의 축복','1','아군전체','ATK버프+SPD버프+정화','ATK+30%(3T), SPD+20(3T), 디버프제거','쿨타임 3턴'],
    ['루미나','Active','빛의 축복+','2','아군전체','ATK버프+SPD버프+정화','ATK+35%(3T), SPD+25(3T), 디버프제거','Lv2 강화'],
    ['루미나','Ultimate','성역의 은총','1','아군전체','힐+ATK+CriDmg+SP','HP35%회복, ATK+35%(3T), CriDmg+50%(3T), SP+1','SP소모: 3'],
    ['루미나','Ultimate','성역의 은총+','2','아군전체','힐+ATK+CriDmg+SP','HP40%회복, ATK+40%(3T), CriDmg+55%(3T), SP+1','SP소모: 3'],
    ['루미나','Ultimate','성역의 은총++','3','아군전체','힐+ATK+CriDmg+SP','HP45%회복, ATK+45%(3T), CriDmg+60%(3T), SP+2','SP소모: 3'],
    ['루미나','Passive','빛의 가호','1','아군전체','전투시작 DEF버프','DEF+15% (영구)','자동 발동'],
    ['루미나','Passive','성스러운 빛','2','피격아군','피격시 확률 회복','HP10% 회복 (5% 확률)','피격 트리거'],
    ['에레보스','Normal','심연의 쌍격','—','적1','2회 연속 공격','ATK×250% × 2타 (실질 500%)','크리 기회 2배'],
    ['에레보스','Active','종결자의 일격','1','적최저HP','처형형 단일 공격+디버프','500% + HP<50%시 +300% + DEF-25%(2T)','쿨타임 4턴'],
    ['에레보스','Active','종결자의 일격+','2','적최저HP','처형형 단일 공격+디버프','600% + HP<50%시 +350% + DEF-30%(2T)','Lv2 강화'],
    ['에레보스','Ultimate','심판의 심연','1','적전체','AoE+디버프+DoT','350% + DEF-30%(2T) + 출혈20%(3T)','SP소모: 5'],
    ['에레보스','Ultimate','심판의 심연+','2','적전체','AoE+디버프+DoT','400% + DEF-35%(2T) + 출혈25%(3T)','SP소모: 5'],
    ['에레보스','Ultimate','심판의 심연++','3','적전체','AoE+디버프+DoT','450% + DEF-40%(2T) + 출혈30%(3T)','SP소모: 5'],
    ['에레보스','Passive','심연의 힘','1','자신','전투시작 스탯버프','ATK+20% + PEN+10% (영구)','자동 발동'],
    ['에레보스','Passive','처형자의 본능','2','자신','적 처치 시 ATK버프','ATK+15% (2턴)','킬 트리거'],
]
for i, row in enumerate(descs, 2):
    for j, val in enumerate(row):
        ws6.cell(row=i, column=j+1, value=val)
    sd(ws6, i, len(dc))
    fc = 'FFF2CC' if '루미나' in str(ws6.cell(row=i,column=1).value) else 'E2EFDA'
    for c in range(1, len(dc)+1):
        ws6.cell(row=i, column=c).fill = PatternFill('solid', fgColor=fc)
print('Sheet 6 done')

# ===== Sheet 7: 밸런스 비교 =====
ws7 = wb.create_sheet('밸런스 비교')
bc = ['항목','루미나(c600)','시트리','엘리시온','브라우니','에레보스(c601)','이브','쿠바바','아르테미스']
for c, name in enumerate(bc, 1):
    ws7.cell(row=1, column=c, value=name)
    ws7.column_dimensions[get_column_letter(c)].width = 16
sh(ws7, 1, len(bc))

bal = [
    ['ATK',230,210,185,195,530,490,430,400],
    ['DEF',160,140,145,150,135,110,120,115],
    ['HP',6800,5600,5800,5400,5500,4800,5000,4900],
    ['SPD',120,110,115,118,80,75,70,78],
    ['CRI','5%','5%','5%','5%','35%','30%','25%','20%'],
    ['PEN','0%','0%','0%','0%','20%','15%','10%','5%'],
    ['Normal배율','120%','100%','90%','95%','250%×2','200%','220%','180%×2'],
    ['Active최대','ATK+35%,SPD+25,정화','ATK+25%','SPD+15','DEF+20%,정화','600%+350%+DEF-30%','400%+HP30%처형','450%','350%×2'],
    ['Ultimate최대','힐45%+ATK45%+CriDmg60%+SP2','ATK+30%+CriDmg30%','힐30%+SP1','DEF+25%+힐20%','450%AoE+DEF-40%+출혈30%','전체280%','전체350%','전체300%+CRI'],
    ['SP소모',3,4,4,3,5,5,6,5],
    ['OP등급','SS+','S','A+','A','SS+','S+','S','A+'],
]
for i, row in enumerate(bal, 2):
    for j, val in enumerate(row):
        ws7.cell(row=i, column=j+1, value=val)
    sd(ws7, i, len(bc))
    ws7.cell(row=i, column=2).fill = PatternFill('solid', fgColor='FFF2CC')
    ws7.cell(row=i, column=6).fill = PatternFill('solid', fgColor='E2EFDA')
ws7.cell(row=1, column=2).fill = PatternFill('solid', fgColor='FFD700')
ws7.cell(row=1, column=6).fill = PatternFill('solid', fgColor='92D050')
print('Sheet 7 done')

out = 'data/Skill_c600_c601_기획서.xlsx'
wb.save(out)
print(f'Saved: {out}')
print(f'Sheets: {wb.sheetnames}')
print(f'Skill rows: {len(c600s)+len(c601s)}')
print(f'BuffValue rows: {len(c600bv)+len(c601bv)}')
print('DONE')
