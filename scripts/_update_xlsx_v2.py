"""Skill.xlsx 패시브 데이터 — 기존 양식에 맞춤 재생성"""
import sys, os, inspect
sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import src.fixtures.test_data as td

wb = load_workbook("data/Skill.xlsx")

cell_font = Font(name="맑은 고딕", size=10)
bold_font = Font(name="맑은 고딕", size=10, bold=True)
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))
center = Alignment(horizontal="center", vertical="center")
left_al = Alignment(horizontal="left", vertical="center", wrap_text=True)

elem_fills = {
    "🔥화": PatternFill("solid", fgColor="FFE0D0"),
    "💧수": PatternFill("solid", fgColor="D0E8FF"),
    "🌿목": PatternFill("solid", fgColor="D8F0D0"),
    "✨광": PatternFill("solid", fgColor="FFF8D0"),
    "🌙암": PatternFill("solid", fgColor="E8D8F0"),
}
event_fills = {
    "on_kill": PatternFill("solid", fgColor="FFDDD0"),
    "on_hit": PatternFill("solid", fgColor="D0DDFF"),
    "on_battle_start": PatternFill("solid", fgColor="D0FFD0"),
}
passive_fill = PatternFill("solid", fgColor="E0FFE0")

ELEM_MAP = {"fire": "🔥화", "water": "💧수", "forest": "🌿목", "light": "✨광", "dark": "🌙암"}

# ═══ 타겟 한국어 변환 (기존 양식) ═══
TARGET_KR = {
    "enemy_near": "적 1인",
    "enemy_lowest_hp": "적 최저HP",
    "enemy_random": "적 랜덤 1인",
    "enemy_random_2": "적 랜덤 2인",
    "enemy_random_3": "적 랜덤 3인",
    "enemy_near_row": "적 행",
    "enemy_near_cross": "적 십자",
    "enemy_back_row": "적 후열",
    "enemy_same_col": "적 같은 열",
    "enemy_highest_spd": "적 최고SPD",
    "all_enemy": "적 전체",
    "self": "자신",
    "ally_lowest_hp": "아군 최저HP",
    "ally_lowest_hp_2": "아군 최저HP 2인",
    "ally_highest_atk": "아군 최고ATK",
    "ally_same_row": "같은 행 아군",
    "ally_dead_random": "사망 아군 1인",
    "all_ally": "아군 전체",
}

# ═══ 트리거 이벤트 한국어 ═══
EVENT_KR = {
    "on_kill": "처치 시",
    "on_hit": "피격 시",
    "on_battle_start": "전투시작",
}

# ═══ Character 시트에서 성급 정보 ═══
char_ws = wb["Character"]
char_info = {}
for row in char_ws.iter_rows(min_row=2, values_only=True):
    if row[0]:
        char_info[str(row[0])] = {"name": row[1], "elem": row[2], "star": row[4]}  # 성급은 5번째 열(index 4)

# ═══ 패시브 캐릭터 수집 ═══
all_chars = []
for name, func in inspect.getmembers(td, inspect.isfunction):
    if name.startswith("make_") and not name.startswith("make_teddy"):
        try:
            c = func()
            if c.passive_skill and c.side == "ally":
                all_chars.append(c)
        except Exception:
            pass
all_chars.sort(key=lambda x: x.id)
print(f"패시브 보유 캐릭터: {len(all_chars)}명")


def get_star(c):
    """성급 문자열 반환"""
    info = char_info.get(c.id, {})
    return info.get("star", "3★")


def describe_effect(eff):
    """이펙트를 기존 한국어 양식으로 변환"""
    lt = eff.logic_type.value
    tgt = TARGET_KR.get(eff.target_type.value, eff.target_type.value)

    if lt == "damage":
        s = f"{tgt} {eff.multiplier:.1f}x 데미지"
        if eff.hit_count > 1:
            s += f"({eff.hit_count}연타)"
        if eff.condition and "burn_bonus_per_stack" in eff.condition:
            s += f"(화상스택당 +{eff.condition['burn_bonus_per_stack']:.2f}x)"
        return s
    elif lt == "heal_hp_ratio":
        return f"{tgt} HP비례 회복 {eff.value*100:.0f}%"
    elif lt == "dot" and eff.buff_data:
        bd = eff.buff_data
        dt_kr = {"burn": "화상", "poison": "중독", "bleed": "출혈"}.get(bd.dot_type, bd.dot_type)
        s = f"{tgt} {dt_kr} {bd.value*100:.0f}% {bd.duration}턴"
        if bd.max_stacks > 1:
            s += f"(최대{bd.max_stacks}스택)"
        return s
    elif lt == "cc" and eff.buff_data:
        bd = eff.buff_data
        cc_kr = {"freeze": "빙결", "stun": "기절", "sleep": "수면", "panic": "공포",
                 "stone": "석화", "electric_shock": "감전"}.get(
            bd.cc_type.value if bd.cc_type else "", "CC")
        return f"{tgt} {cc_kr} {bd.duration}턴"
    elif lt == "stat_change" and eff.buff_data:
        bd = eff.buff_data
        stat_kr = {"atk": "공격력", "def_": "방어력", "spd": "속도",
                   "cri_ratio": "크리율", "cri_dmg_ratio": "크리피해",
                   "acc": "명중"}.get(bd.stat, bd.stat or "")
        if bd.is_ratio:
            val_str = f"{bd.value*100:.0f}%"
        else:
            val_str = f"{bd.value:.0f}"
        sign = "-" if bd.is_debuff else "+"
        return f"{tgt} {stat_kr}{sign}{val_str} {bd.duration}턴"
    elif lt == "barrier_ratio":
        return f"{tgt} 보호막 {eff.value*100:.0f}%"
    elif lt == "barrier":
        return f"{tgt} 보호막"
    elif lt == "buff_turn_increase":
        return f"{tgt} 버프턴+{eff.value:.0f}"
    elif lt == "debuff_turn_increase":
        return f"{tgt} 디버프턴+{eff.value:.0f}"
    elif lt == "remove_debuff":
        return f"{tgt} 디버프제거"
    elif lt == "remove_buff":
        return f"{tgt} 버프제거"
    elif lt == "sp_steal":
        return f"{tgt} SP강탈 {eff.value:.0f}"
    elif lt == "sp_lock":
        return f"{tgt} SP잠금 {eff.value:.0f}턴"
    elif lt == "damage_penetration":
        return f"{tgt} 관통 {eff.multiplier:.1f}x"
    elif lt == "damage_cri":
        return f"{tgt} 확정크리 {eff.multiplier:.1f}x"
    elif lt == "damage_hp_ratio":
        return f"{tgt} 최대HP {eff.value*100:.0f}% 데미지"
    elif lt == "debuff_immune":
        return f"{tgt} 디버프면역 {eff.value:.0f}턴"
    elif lt == "invincibility":
        return f"{tgt} 무적 {eff.value:.0f}턴"
    elif lt == "undying":
        return f"{tgt} 불사 {eff.value:.0f}턴"
    elif lt == "dot_heal_hp_ratio" and eff.buff_data:
        bd = eff.buff_data
        return f"{tgt} 재생 {bd.value*100:.0f}% {bd.duration}턴"
    elif lt == "ignore_element":
        return f"{tgt} 속성무시 {eff.value:.0f}턴"
    elif lt == "cri_unavailable":
        return f"{tgt} 크리불가 {eff.value:.0f}턴"
    elif lt == "counter_unavailable":
        return f"{tgt} 반격불가 {eff.value:.0f}턴"
    elif lt == "active_cd_change":
        return f"{tgt} 쿨타임{eff.value:+.0f}"
    elif lt == "revive":
        return f"{tgt} 부활 HP{eff.value*100:.0f}%"
    elif lt == "taunt":
        return f"{tgt} 도발 {eff.value:.0f}턴"
    elif lt == "counter":
        return f"{tgt} 반격 {eff.value:.0f}턴"
    elif lt == "sp_increase":
        return f"{tgt} SP+{eff.value:.0f}"
    else:
        return lt


def make_passive_desc(c):
    """패시브 설명을 기존 양식으로 생성"""
    ps = c.passive_skill
    trigger_desc = ""
    for t in c.triggers:
        ev = EVENT_KR.get(t.event.value, t.event.value)
        once = ", 1회" if t.once_per_battle else ""
        trigger_desc = f"[{ev}{once}] "

    effs = [describe_effect(eff) for eff in ps.effects]
    return trigger_desc + ", ".join(effs)


# ═══ 1. Skill 시트: 기존 passive 행 제거 후 재추가 ═══
skill_ws = wb["Skill"]
# 기존 passive 행 제거 (뒤에서부터)
rows_to_delete = []
for row in range(skill_ws.max_row, 1, -1):
    if skill_ws.cell(row=row, column=7).value == "passive":
        rows_to_delete.append(row)
for r in rows_to_delete:
    skill_ws.delete_rows(r)
print(f"기존 passive 행 {len(rows_to_delete)}개 제거")

skill_row = skill_ws.max_row + 1
passive_count = 0

for c in all_chars:
    ps = c.passive_skill
    elem = ELEM_MAP.get(c.element.value, c.element.value)
    star = get_star(c)

    for eff_idx, eff in enumerate(ps.effects, 1):
        vals = [
            c.id, c.name, elem, star, ps.id, ps.name, "passive", 0, 0, eff_idx,
            eff.logic_type.value, eff.target_type.value, eff.value, eff.multiplier, eff.hit_count,
        ]
        if eff.buff_data:
            bd = eff.buff_data
            vals += [bd.id, bd.name, bd.logic_type.value, bd.stat, bd.value,
                     "Y" if bd.is_ratio else "N", bd.duration,
                     "Y" if bd.is_debuff else "N",
                     bd.cc_type.value if bd.cc_type else None,
                     bd.dot_type,
                     bd.max_stacks if bd.max_stacks > 1 else None]
        else:
            vals += [None] * 11

        for j, val in enumerate(vals):
            cell = skill_ws.cell(row=skill_row, column=j + 1, value=val)
            cell.font = cell_font
            cell.border = thin_border
            cell.alignment = center if j not in (5, 10, 11) else left_al
            if j == 2 and val in elem_fills:
                cell.fill = elem_fills[val]
            if j == 6:
                cell.fill = passive_fill
        skill_row += 1
    passive_count += 1

print(f"Skill 시트: {passive_count}개 패시브 스킬 추가")

# ═══ 2. Trigger 시트: 전체 재작성 ═══
trig_ws = wb["Trigger"]
for row in range(trig_ws.max_row, 1, -1):
    trig_ws.delete_rows(row)

trig_row = 2
for c in all_chars:
    elem = ELEM_MAP.get(c.element.value, c.element.value)
    star = get_star(c)
    for trigger in c.triggers:
        vals = [c.id, c.name, elem, star, trigger.event.value, None,
                trigger.skill_id, trigger.buff_id,
                "Y" if trigger.once_per_battle else "N"]
        for j, val in enumerate(vals):
            cell = trig_ws.cell(row=trig_row, column=j + 1, value=val)
            cell.font = cell_font
            cell.border = thin_border
            cell.alignment = center
            if j == 2 and val in elem_fills:
                cell.fill = elem_fills[val]
            if j == 4 and val in event_fills:
                cell.fill = event_fills[val]
        trig_row += 1

print(f"Trigger 시트: {trig_row - 2}개 트리거 재작성")

# ═══ 3. Character 시트: Passive이름/설명 업데이트 ═══
header_vals = [char_ws.cell(row=1, column=c).value for c in range(1, char_ws.max_column + 1)]
if "Passive이름" in header_vals:
    pcol_name = header_vals.index("Passive이름") + 1
    pcol_desc = pcol_name + 1
else:
    pcol_name = char_ws.max_column + 1
    pcol_desc = pcol_name + 1
    char_ws.cell(row=1, column=pcol_name, value="Passive이름").font = bold_font
    char_ws.cell(row=1, column=pcol_name).border = thin_border
    char_ws.cell(row=1, column=pcol_name).alignment = center
    char_ws.cell(row=1, column=pcol_desc, value="Passive설명").font = bold_font
    char_ws.cell(row=1, column=pcol_desc).border = thin_border
    char_ws.cell(row=1, column=pcol_desc).alignment = center

passive_map = {}
for c in all_chars:
    passive_map[c.id] = (c.passive_skill.name, make_passive_desc(c))

for row in range(2, char_ws.max_row + 1):
    cid = str(char_ws.cell(row=row, column=1).value)
    if cid in passive_map:
        pname, pdesc = passive_map[cid]
        cell_n = char_ws.cell(row=row, column=pcol_name, value=pname)
        cell_n.font = cell_font
        cell_n.border = thin_border
        cell_n.alignment = center
        cell_d = char_ws.cell(row=row, column=pcol_desc, value=pdesc)
        cell_d.font = cell_font
        cell_d.border = thin_border
        cell_d.alignment = left_al

# 열 너비
from openpyxl.utils import get_column_letter
char_ws.column_dimensions[get_column_letter(pcol_name)].width = 14
char_ws.column_dimensions[get_column_letter(pcol_desc)].width = 55

print(f"Character 시트: Passive이름/설명 업데이트")

try:
    wb.save("data/Skill.xlsx")
    print("✅ data/Skill.xlsx 저장 완료")
except PermissionError:
    wb.save("data/Skill_v2.xlsx")
    print("⚠️  Skill.xlsx 열려있어 data/Skill_v2.xlsx로 저장")
