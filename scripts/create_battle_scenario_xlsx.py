"""M04 vs M01 배틀 시나리오 스프레드시트 생성"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
thin = Side(style='thin')
border = Border(top=thin, bottom=thin, left=thin, right=thin)
header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='2F5496')
data_font = Font(name='Arial', size=10)
kill_fill = PatternFill('solid', fgColor='FFD9D9')
m04_fill = PatternFill('solid', fgColor='E8D5F5')
m01_fill = PatternFill('solid', fgColor='FCE4D6')
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border


def style_data(ws, row, cols, fill=None):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = data_font
        cell.alignment = center if c < cols else left_wrap
        cell.border = border
        if fill:
            cell.fill = fill


# ═══ Sheet 1: 교전 정보 ═══
ws1 = wb.active
ws1.title = '교전 정보'
ws1.sheet_properties.tabColor = '2F5496'

ws1.merge_cells('A1:J1')
ws1['A1'] = 'M04 하이퍼캐리 vs M01 화상지옥 - 교전 정보'
ws1['A1'].font = Font(name='Arial', bold=True, size=14, color='2F5496')
ws1['A1'].alignment = Alignment(horizontal='center')

headers = ['진영', '캐릭터', '역할', '속성', 'ATK', 'DEF', 'HP', 'SPD', 'SP비용', '핵심 스킬']
for i, h in enumerate(headers, 1):
    ws1.cell(row=3, column=i, value=h)
style_header(ws1, 3, 10)

m04_data = [
    ['M04', '쿠바바', 'ATTACKER', '암', 400, 200, 4000, 80, 6, 'Ult 3.50x 단일 + ON_KILL Normal 추격'],
    ['M04', '루미나', 'MAGICIAN', '광', 345, 230, 5750, 100, 4, 'ATK+25% + CRI_DMG+0.30 + 힐25% + 정화'],
    ['M04', '시트리', 'SUPPORTER', '광', 250, 200, 5000, 120, 4, 'ATK+25% + CRI_DMG+0.20 + SPD+10'],
    ['M04', '티타니아', 'DEFENDER', '광', 250, 300, 6000, 110, 3, '힐20% + ATK+20% + SPD+10 + CRI+15%'],
    ['M04', '프레이야', 'HEALER', '목', 120, 120, 2400, 90, 3, '힐35% + DEF+15% + 정화'],
]
m01_data = [
    ['M01', '모건', 'MAGICIAN', '화', 300, 200, 5000, 100, 4, 'Normal 0.90x x2연타 + Ult AoE 1.60x + 출혈'],
    ['M01', '다비', 'MAGICIAN', '화', 300, 200, 5000, 100, 4, 'Active 화상 + Ult AoE 1.40x + 화상'],
    ['M01', '카라라트리', 'ATTACKER', '화', 400, 200, 4000, 80, 6, 'Normal burn_bonus 0.40x/스택 + Ult DEF-20%'],
    ['M01', '드미테르', 'SUPPORTER', '화', 200, 160, 4000, 120, 4, 'AoE 화상20% + SPD-15% + burn_bonus 0.20'],
    ['M01', '지바', 'HEALER', '화', 200, 200, 4000, 90, 3, '힐30% + ATK+20% + SP+1'],
]

for r, row in enumerate(m04_data, 4):
    for c, v in enumerate(row, 1):
        ws1.cell(row=r, column=c, value=v)
    style_data(ws1, r, 10, m04_fill)

for r, row in enumerate(m01_data, 9):
    for c, v in enumerate(row, 1):
        ws1.cell(row=r, column=c, value=v)
    style_data(ws1, r, 10, m01_fill)

ws1.merge_cells('A15:J15')
c15 = ws1['A15']
c15.value = 'RESULT: M04 승리 | 240턴 / 89.2초 / 8라운드 | M04 생존: 루미나/시트리/티타니아/프레이야(100%) | 쿠바바 전사(T31)'
c15.font = Font(name='Arial', bold=True, size=11, color='2F5496')
c15.alignment = Alignment(horizontal='center', wrap_text=True)
ws1.row_dimensions[15].height = 30

for c in [1, 2, 3, 4]:
    ws1.column_dimensions[get_column_letter(c)].width = 12
for c in [5, 6, 7, 8, 9]:
    ws1.column_dimensions[get_column_letter(c)].width = 8
ws1.column_dimensions['J'].width = 40

# ═══ Sheet 2: 턴별 타임라인 ═══
ws2 = wb.create_sheet('턴별 타임라인')
ws2.sheet_properties.tabColor = 'E67E22'

ws2.merge_cells('A1:L1')
ws2['A1'] = '턴별 타임라인 - 주요 이벤트'
ws2['A1'].font = Font(name='Arial', bold=True, size=14, color='E67E22')
ws2['A1'].alignment = Alignment(horizontal='center')

h2 = ['턴', '시간', 'R', 'SP(M04)', 'SP(M01)', '행동자', '진영', '스킬', '타입', '대상', '피해량', '비고']
for i, h in enumerate(h2, 1):
    ws2.cell(row=3, column=i, value=h)
style_header(ws2, 3, 12)

timeline = [
    ['T1', 2.50, 1, 1, 1, '시트리', 'M04', '크런치 캔디', 'Active', '아군 전체', '-', '오프닝 ATK+25% SPD+10'],
    ['T2', 2.50, 1, 2, 2, '드미테르', 'M01', '사이드 이펙트', 'Active', '적 전체', '50~75', '화상20% + SPD-15% 감속'],
    ['T3', 2.68, 1, 3, 3, '티타니아', 'M04', '요정의 여왕', 'Ult(SP:3)', '아군 전체', '-', '첫 Ult! 힐+ATK+SPD 버프'],
    ['T5', 2.88, 1, 1, 4, '루미나', 'M04', '빛의 축복', 'Active', '아군 전체', '-', 'ATK+15% SPD+15 + 감속 정화!'],
    ['T7', 3.00, 1, '-', 1, '모건', 'M01', '화도난무', 'Ult(SP:4)', '적 전체', '349~737', 'AoE+출혈, 프레이야 위기!'],
    ['T9', 3.10, 1, '-', '-', '프레이야', 'M04', '대지의 은총', 'Ult(SP:3)', '아군 전체', '-', '긴급 힐35% + DEF+15%'],
    ['T12', 3.33, 1, '-', 1, '지바', 'M01', '대자연의 손길', 'Ult(SP:3)', '아군 전체', '-', '힐30% + ATK+20% + SP+1'],
    ['T13', 3.35, 1, 3, 7, '쿠바바', 'M04', '오버 더 리밋', 'Active', '적 전체', '877~1316', '*** 트리플 크리! DEF-20%'],
    ['T15', 3.75, 1, '-', '-', '카라라트리', 'M01', '출세간의 선법', 'Ult(SP:6)', '적 전체', '363~603', 'AoE + DEF-20% 반격'],
    ['T17', 4.28, 1, '-', '-', '시트리', 'M04', '라 돌체 비타', 'Ult(SP:4)', '전체', '431~466', 'ATK+25% CRI_DMG+0.20 적층'],
    ['T22', 5.11, 1, '-', '-', '루미나', 'M04', '성역의 은총', 'Ult(SP:4)', '아군 전체', '-', '최종 버프! ATK+25% CRI_DMG+0.30'],
    ['T24', 5.96, 1, '-', '-', '쿠바바', 'M04', '와일드 로드', 'Normal', '드미테르', '1,299', '드미테르 빈사! (HP 321)'],
    ['T28', 6.21, 1, '-', '-', '시트리', 'M04', '폴 인 러브', 'Normal', '드미테르', '768 CRI', 'KILL: 드미테르 사망 (1st)'],
    ['T31', 7.50, 1, '-', '-', '카라라트리', 'M01', '무기법', 'Normal', '쿠바바', 'burn 2.30x', 'KILL: 쿠바바 사망 (burn_bonus)'],
    ['T132', 42.2, 4, '-', '-', '루미나', 'M04', 'Normal', 'Normal', '다비', '-', 'KILL: 다비 사망 (소모전)'],
    ['T160', 52.7, 5, '-', '-', '루미나', 'M04', 'Normal', 'Normal', '모건', '-', 'KILL: 모건 사망'],
    ['T203', 72.3, 7, '-', '-', '티타니아', 'M04', 'Normal', 'Normal', '카라라트리', '-', 'KILL: 카라라트리 사망'],
    ['T240', 89.2, 8, '-', '-', '티타니아', 'M04', 'Normal', 'Normal', '지바', '-', 'KILL: 지바 사망 -> M04 승리!'],
]

for r, row in enumerate(timeline, 4):
    fill = None
    if 'KILL' in str(row[-1]):
        fill = kill_fill
    elif row[6] == 'M04':
        fill = PatternFill('solid', fgColor='F2E6FF')
    else:
        fill = PatternFill('solid', fgColor='FFF2E6')
    for c, v in enumerate(row, 1):
        ws2.cell(row=r, column=c, value=v)
    style_data(ws2, r, 12, fill)

ws2.column_dimensions['A'].width = 6
ws2.column_dimensions['B'].width = 7
ws2.column_dimensions['C'].width = 4
ws2.column_dimensions['D'].width = 9
ws2.column_dimensions['E'].width = 9
ws2.column_dimensions['F'].width = 10
ws2.column_dimensions['G'].width = 6
ws2.column_dimensions['H'].width = 14
ws2.column_dimensions['I'].width = 10
ws2.column_dimensions['J'].width = 10
ws2.column_dimensions['K'].width = 11
ws2.column_dimensions['L'].width = 34

# ═══ Sheet 3: 사망 순서 ═══
ws3 = wb.create_sheet('사망 순서')
ws3.sheet_properties.tabColor = 'C0392B'

ws3.merge_cells('A1:H1')
ws3['A1'] = '사망 순서 & 전투 결과'
ws3['A1'].font = Font(name='Arial', bold=True, size=14, color='C0392B')

h3 = ['순서', '턴', '시간', '사망 캐릭터', '진영', '처치자', '처치 스킬', '사인']
for i, h in enumerate(h3, 1):
    ws3.cell(row=3, column=i, value=h)
style_header(ws3, 3, 8)

deaths = [
    [1, 'T28', '6.21초', '드미테르', 'M01', '시트리', 'Normal CRI 768', 'M01 화상 핵심 + 감속 담당 탈락'],
    [2, 'T31', '7.50초', '쿠바바', 'M04', '카라라트리', 'burn_bonus 2.30x', 'M04 에이스 전사 (화상2스택 폭딜)'],
    [3, 'T132', '42.2초', '다비', 'M01', '루미나', 'Normal 지속 공격', 'M01 화상 소스 2명->1명 감소'],
    [4, 'T160', '52.7초', '모건', 'M01', '루미나', 'Normal 지속 공격', 'M01 출혈 소스 완전 탈락'],
    [5, 'T203', '72.3초', '카라라트리', 'M01', '티타니아', 'Normal 소모전', 'burn_bonus 주역 탈락'],
    [6, 'T240', '89.2초', '지바', 'M01', '티타니아', 'Normal 마무리', 'M01 전멸, M04 승리'],
]

for r, row in enumerate(deaths, 4):
    for c, v in enumerate(row, 1):
        ws3.cell(row=r, column=c, value=v)
    style_data(ws3, r, 8, kill_fill)

widths3 = [6, 6, 8, 10, 6, 10, 18, 34]
for c in range(1, 9):
    ws3.column_dimensions[get_column_letter(c)].width = widths3[c - 1]

# ═══ Sheet 4: SP 타이밍 ═══
ws4 = wb.create_sheet('SP 타이밍')
ws4.sheet_properties.tabColor = '27AE60'

ws4.merge_cells('A1:H1')
ws4['A1'] = 'SP 타이밍 & 얼티밋 발동 순서'
ws4['A1'].font = Font(name='Arial', bold=True, size=14, color='27AE60')

h4 = ['순서', '턴', '진영', '캐릭터', 'SP비용', '얼티밋명', '효과', '전략적 의미']
for i, h in enumerate(h4, 1):
    ws4.cell(row=3, column=i, value=h)
style_header(ws4, 3, 8)

sp_data = [
    [1, 'T3', 'M04', '티타니아', 3, '요정의 여왕', '힐20%+ATK+20%+SPD+10', '첫 Ult로 버프 선점'],
    [2, 'T7', 'M01', '모건', 4, '화도난무', '1.60x AoE + 출혈', 'M04에 DoT 압력'],
    [3, 'T9', 'M04', '프레이야', 3, '대지의 은총', '힐35% + DEF+15%', '긴급 방어'],
    [4, 'T12', 'M01', '지바', 3, '대자연의 손길', '힐30%+ATK+20%+SP+1', 'M01 안정화+SP 가속'],
    [5, 'T15', 'M01', '카라라트리', 6, '출세간의 선법', '1.30x AoE+DEF-20%', 'M04 방어선 약화'],
    [6, 'T17', 'M04', '시트리', 4, '라 돌체 비타', 'ATK+25%+CRI_DMG+0.20', '쿠바바 풀버프 완성'],
    [7, 'T20', 'M01', '드미테르', 4, '극약처방', '2.20x+DEF-20%', '쿠바바 집중 약화'],
    [8, 'T22', 'M04', '루미나', 4, '성역의 은총', '힐25%+ATK+25%+CRI_DMG+0.30', '최종 버프 적층'],
]

for r, row in enumerate(sp_data, 4):
    fill = m04_fill if row[2] == 'M04' else m01_fill
    for c, v in enumerate(row, 1):
        ws4.cell(row=r, column=c, value=v)
    style_data(ws4, r, 8, fill)

widths4 = [6, 6, 6, 10, 7, 14, 26, 30]
for c in range(1, 9):
    ws4.column_dimensions[get_column_letter(c)].width = widths4[c - 1]

# ═══ Sheet 5: 콤보 발동 기록 ═══
ws5 = wb.create_sheet('콤보 발동 기록')
ws5.sheet_properties.tabColor = '8E44AD'

ws5.merge_cells('A1:F1')
ws5['A1'] = 'v6.0 콤보 메커니즘 발동 기록'
ws5['A1'].font = Font(name='Arial', bold=True, size=14, color='8E44AD')

h5 = ['메커니즘', '캐릭터', '발동 턴', '조건', '결과', '전투 영향']
for i, h in enumerate(h5, 1):
    ws5.cell(row=3, column=i, value=h)
style_header(ws5, 3, 6)

combo_data = [
    ['ON_KILL 추격', '쿠바바', '미발동', 'T31 사망으로 킬 기회 없음', '추격 0회', 'M04 에이스가 킬 전에 사망'],
    ['burn_bonus', '카라라트리', 'T31', '화상 2스택', '1.50+(0.40x2)=2.30x', '*** 쿠바바 처치! M01 유일한 킬'],
    ['burn_bonus', '드미테르', 'T2~', '자체 화상 부여 후', '0.50+(0.20xN)x', 'AoE 칩데미지 소폭 증가'],
    ['SPD 감속', '드미테르', 'T2', 'Active 발동', 'M04 전원 SPD-15%', 'T5 루미나 정화로 2턴 무효화'],
    ['디버프 정화', '루미나', 'T5', 'Active 발동', '화상 감속 해제', 'M01 감속 전략 완전 차단'],
    ['2연타(hit_count)', '모건', '매 Normal', 'Normal 0.90x x2', '출혈 2회 적용 기회', 'DoT 중첩 속도 향상'],
    ['ON_HIT 반격', '-', '-', '해당 캐릭터 없음', '-', '이 매치업에서 미사용'],
    ['ON_BATTLE_START', '-', '-', '시트리 트리거 제거', '-', 'v6.0 밸런스 조정으로 비활성화'],
]

highlight = PatternFill('solid', fgColor='F5EEF8')
for r, row in enumerate(combo_data, 4):
    for c, v in enumerate(row, 1):
        ws5.cell(row=r, column=c, value=v)
    fill = highlight if '***' in str(row[5]) else None
    style_data(ws5, r, 6, fill)

widths5 = [16, 10, 8, 24, 22, 36]
for c in range(1, 7):
    ws5.column_dimensions[get_column_letter(c)].width = widths5[c - 1]

# Save
out = 'C:/Ai/BattleAgent/docs/M04_vs_M01_배틀시나리오.xlsx'
wb.save(out)
print(f'Saved: {out}')
print(f'Sheets: {wb.sheetnames}')
