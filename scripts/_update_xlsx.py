"""Skill.xlsx에 패시브 스킬 데이터 + Trigger 시트 업데이트"""
import sys, os, inspect
sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import src.fixtures.test_data as td

wb = load_workbook("data/Skill.xlsx")

cell_font = Font(name="맑은 고딕", size=10)
bold_font = Font(name="맑은 고딕", size=10, bold=True)
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))
center = Alignment(horizontal="center", vertical="center")
left_al = Alignment(horizontal="left", vertical="center")

elem_fills = {
    "🔥화": PatternFill("solid", fgColor="FFE0D0"),
    "💧수": PatternFill("solid", fgColor="D0E8FF"),
    "🌿목": PatternFill("solid", fgColor="D8F0D0"),
    "✨광": PatternFill("solid", fgColor="FFF8D0"),
    "🌙암": PatternFill("solid", fgColor="E8D8F0"),
}
event_fills = {
    "on_kill": PatternFill("solid", fgColor="FFDDD0"),
    "on_hit": PatternFill("solid", fgColor="D0DDFF"),
    "on_battle_start": PatternFill("solid", fgColor="D0FFD0"),
}
passive_fill = PatternFill("solid", fgColor="E0FFE0")

ELEM_MAP = {"fire": "🔥화", "water": "💧수", "forest": "🌿목", "light": "✨광", "dark": "🌙암"}

# Character 시트에서 성급 정보
char_ws = wb["Character"]
char_info = {}
for row in char_ws.iter_rows(min_row=2, values_only=True):
    if row[0]:
        char_info[str(row[0])] = {"name": row[1], "elem": row[2], "star": row[3]}

# 패시브 보유 캐릭터 수집
all_chars = []
for name, func in inspect.getmembers(td, inspect.isfunction):
    if name.startswith("make_") and not name.startswith("make_teddy"):
        try:
            c = func()
            if c.passive_skill and c.side == "ally":
                all_chars.append(c)
        except Exception:
            pass

all_chars.sort(key=lambda x: x.id)
print(f"패시브 보유 캐릭터: {len(all_chars)}명")

# ═══ Skill 시트에 패시브 행 추가 ═══
skill_ws = wb["Skill"]
skill_row = skill_ws.max_row + 1
passive_count = 0

for c in all_chars:
    ps = c.passive_skill
    elem = ELEM_MAP.get(c.element.value, c.element.value)
    star = char_info.get(c.id, {}).get("star", "3★")

    for eff_idx, eff in enumerate(ps.effects, 1):
        vals = [
            c.id, c.name, elem, star, ps.id, ps.name, "passive", 0, 0, eff_idx,
            eff.logic_type.value, eff.target_type.value, eff.value, eff.multiplier, eff.hit_count,
        ]
        if eff.buff_data:
            bd = eff.buff_data
            vals += [bd.id, bd.name, bd.logic_type.value, bd.stat, bd.value,
                     "Y" if bd.is_ratio else "N", bd.duration,
                     "Y" if bd.is_debuff else "N",
                     bd.cc_type.value if bd.cc_type else None,
                     bd.dot_type,
                     bd.max_stacks if bd.max_stacks > 1 else None]
        else:
            vals += [None]*11

        for j, val in enumerate(vals):
            cell = skill_ws.cell(row=skill_row, column=j+1, value=val)
            cell.font = cell_font
            cell.border = thin_border
            cell.alignment = center
            if j == 2 and val in elem_fills:
                cell.fill = elem_fills[val]
            if j == 6:
                cell.fill = passive_fill

        skill_row += 1
    passive_count += 1

print(f"Skill 시트: {passive_count}개 패시브 스킬 추가")

# ═══ Trigger 시트 전체 재작성 ═══
trig_ws = wb["Trigger"]
for row in range(trig_ws.max_row, 1, -1):
    trig_ws.delete_rows(row)

trig_row = 2
for c in all_chars:
    elem = ELEM_MAP.get(c.element.value, c.element.value)
    star = char_info.get(c.id, {}).get("star", "3★")
    for trigger in c.triggers:
        vals = [c.id, c.name, elem, star, trigger.event.value, None,
                trigger.skill_id, trigger.buff_id,
                "Y" if trigger.once_per_battle else "N"]
        for j, val in enumerate(vals):
            cell = trig_ws.cell(row=trig_row, column=j+1, value=val)
            cell.font = cell_font
            cell.border = thin_border
            cell.alignment = center
            if j == 2 and val in elem_fills:
                cell.fill = elem_fills[val]
            if j == 4 and val in event_fills:
                cell.fill = event_fills[val]
        trig_row += 1

print(f"Trigger 시트: {trig_row - 2}개 트리거 재작성")

# ═══ Character 시트에 Passive 컬럼 추가 ═══
last_col = char_ws.max_column
# 이미 Passive이름 컬럼이 있으면 스킵
header_vals = [char_ws.cell(row=1, column=c).value for c in range(1, last_col+1)]
if "Passive이름" not in header_vals:
    pcol_name = last_col + 1
    pcol_desc = last_col + 2
    char_ws.cell(row=1, column=pcol_name, value="Passive이름").font = bold_font
    char_ws.cell(row=1, column=pcol_name).border = thin_border
    char_ws.cell(row=1, column=pcol_name).alignment = center
    char_ws.cell(row=1, column=pcol_desc, value="Passive설명").font = bold_font
    char_ws.cell(row=1, column=pcol_desc).border = thin_border
    char_ws.cell(row=1, column=pcol_desc).alignment = center
else:
    pcol_name = header_vals.index("Passive이름") + 1
    pcol_desc = pcol_name + 1

# 패시브 설명 생성
passive_map = {}
for c in all_chars:
    ps = c.passive_skill
    trigger_desc = ""
    for t in c.triggers:
        ev = t.event.value.replace("on_", "").replace("_", " ")
        once = " (1회)" if t.once_per_battle else ""
        trigger_desc = f"[{ev}{once}] "

    effs = []
    for eff in ps.effects:
        lt = eff.logic_type.value
        tgt = eff.target_type.value.replace("_", " ")
        if lt == "damage":
            bbs = ""
            if eff.condition and "burn_bonus_per_stack" in eff.condition:
                bbs = f"+burn_bonus {eff.condition['burn_bonus_per_stack']}"
            effs.append(f"{tgt} {eff.multiplier}x{bbs}")
        elif lt == "heal_hp_ratio":
            effs.append(f"{tgt} HP{eff.value*100:.0f}%회복")
        elif lt == "dot":
            dt = eff.buff_data.dot_type if eff.buff_data else "dot"
            effs.append(f"{tgt} {dt} {eff.buff_data.value*100:.0f}%")
        elif lt == "cc":
            ct = eff.buff_data.cc_type.value if eff.buff_data and eff.buff_data.cc_type else "cc"
            effs.append(f"{tgt} {ct}")
        elif lt == "stat_change" and eff.buff_data:
            stat = eff.buff_data.stat or ""
            val = eff.buff_data.value
            pct = "%" if eff.buff_data.is_ratio else ""
            sign = "+" if not eff.buff_data.is_debuff else "-"
            effs.append(f"{tgt} {stat}{sign}{abs(val)*100:.0f}{pct}")
        elif lt == "barrier_ratio":
            effs.append(f"{tgt} 보호막{eff.value*100:.0f}%")
        elif lt == "barrier":
            effs.append(f"{tgt} 보호막")
        elif lt == "buff_turn_increase":
            effs.append(f"{tgt} 버프턴+{eff.value:.0f}")
        elif lt == "debuff_turn_increase":
            effs.append(f"{tgt} 디버프턴+{eff.value:.0f}")
        elif lt == "remove_debuff":
            effs.append(f"{tgt} 디버프제거")
        elif lt == "remove_buff":
            effs.append(f"{tgt} 버프제거")
        elif lt == "sp_steal":
            effs.append(f"{tgt} SP강탈{eff.value:.0f}")
        elif lt == "damage_penetration":
            effs.append(f"{tgt} 관통{eff.multiplier}x")
        elif lt == "damage_cri":
            effs.append(f"{tgt} 확정크리{eff.multiplier}x")
        elif lt == "debuff_immune":
            effs.append(f"{tgt} 디버프면역")
        elif lt == "dot_heal_hp_ratio" and eff.buff_data:
            effs.append(f"{tgt} HoT{eff.buff_data.value*100:.0f}%")
        else:
            effs.append(lt)

    passive_map[c.id] = (ps.name, trigger_desc + ", ".join(effs))

for row in range(2, char_ws.max_row + 1):
    cid = str(char_ws.cell(row=row, column=1).value)
    if cid in passive_map:
        pname, pdesc = passive_map[cid]
        cell_n = char_ws.cell(row=row, column=pcol_name, value=pname)
        cell_n.font = cell_font
        cell_n.border = thin_border
        cell_n.alignment = center
        cell_d = char_ws.cell(row=row, column=pcol_desc, value=pdesc)
        cell_d.font = cell_font
        cell_d.border = thin_border
        cell_d.alignment = left_al

print(f"Character 시트: Passive이름/설명 컬럼 추가")

try:
    wb.save("data/Skill.xlsx")
except PermissionError:
    wb.save("data/Skill_v2.xlsx")
    print("⚠️  Skill.xlsx가 열려있어 Skill_v2.xlsx로 저장")
print("✅ Skill.xlsx 저장 완료")
