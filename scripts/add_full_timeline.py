"""풀 타임라인 시트를 배틀시나리오 xlsx에 추가"""
import sys, re
sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, 'src')

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from battle.battle_engine import BattleEngine
from fixtures.test_data import (
    make_kubaba, make_c600, make_sitri, make_titania, make_freya,
    make_morgan, make_dabi, make_kararatri, make_demeter, make_jiva,
)

# Run battle
ally = [make_kubaba(), make_c600(), make_sitri(), make_titania(), make_freya()]
enemy_chars = [make_morgan(), make_dabi(), make_kararatri(), make_demeter(), make_jiva()]
for c in enemy_chars:
    c.side = 'enemy'

engine = BattleEngine(ally, enemy_chars, seed=42, deck_type='offense')
engine.run()
log = engine.get_log()

# Parse log into structured rows
turn_re = re.compile(r'\[T(\d+)\] t=([\d.]+)(\s*\[Extra Turn\])?\s*\|\s*(.+?)\s*\((\w+)\)\s*\|\s*HP\s*([\d.]+)/([\d.]+)\s*\|\s*SP 아군:(\d+)\s*적:(\d+)')
dmg_re = re.compile(r'(\S+)\s*→\s*(\S+):\s*(.+?)\s+(\d+)\s*피해(\s*\[크리티컬!\])?\s*\(HP\s*([\d.]+)/([\d.]+)\)')
heal_re = re.compile(r'(\S+)\s*→\s*(\S+):\s*(.+?)\s+HP\s*[\d.]+%\s*회복\s*\((\d+),\s*HP\s*([\d.]+)/([\d.]+)\)')
dot_re = re.compile(r'→\s*(\S+)\s*(burn|poison|bleed)\s*피해\s*(\d+)')
buff_re = re.compile(r'→\s*(\S+)\s*버프 추가:\s*(.+?)\s*\((\d+)턴\)')
cc_re = re.compile(r'(\S+)\s*→\s*(\S+):\s*(\w+)\s*부여\s*\((\d+)턴\)')
kill_re = re.compile(r'💀\s*(\S+)\s*사망')
ult_re = re.compile(r'💥\s*(\S+)\s*얼티밋:\s*(.+)')
ult_reserve_re = re.compile(r'✨\s*(\S+)\s*얼티밋 예약!\s*SP\s*-(\d+)')
round_re = re.compile(r'🔔\s*배틀 라운드\s*(\d+)')
trigger_re = re.compile(r'⚡\s*(\S+)\s*트리거 발동')
hard_cc_re = re.compile(r'⛔\s*(\S+)\s*.+로 행동 (불가|실패)')

rows = []
current_turn = None
current_data = None

for line in log:
    m = turn_re.search(line)
    if m:
        if current_data:
            rows.append(current_data)
        turn_num = int(m.group(1))
        time_val = float(m.group(2))
        is_extra = bool(m.group(3))
        char_name = m.group(4)
        side = m.group(5)
        hp_cur = float(m.group(6))
        hp_max = float(m.group(7))
        sp_ally = int(m.group(8))
        sp_enemy = int(m.group(9))
        current_data = {
            'turn': turn_num, 'time': time_val, 'extra': is_extra,
            'char': char_name, 'side': side,
            'hp': f'{hp_cur:.0f}/{hp_max:.0f}',
            'sp_ally': sp_ally, 'sp_enemy': sp_enemy,
            'skill': '', 'type': '', 'actions': [], 'events': []
        }
        continue

    if current_data is None:
        # Check round change
        rm = round_re.search(line)
        if rm:
            rows.append({
                'turn': '-', 'time': '-', 'extra': False,
                'char': '-', 'side': '-', 'hp': '-',
                'sp_ally': 0, 'sp_enemy': 0,
                'skill': f'라운드 {rm.group(1)} 시작', 'type': 'ROUND',
                'actions': [], 'events': ['SP 초기화']
            })
        continue

    # Skill used
    if '🗡' in line:
        parts = line.split('→')
        if len(parts) > 1:
            sk = parts[-1].strip().split('(')
            current_data['skill'] = sk[0].strip()
            current_data['type'] = sk[1].replace(')', '').strip() if len(sk) > 1 else 'normal'

    # Ultimate
    um = ult_re.search(line)
    if um:
        current_data['skill'] = um.group(2)
        current_data['type'] = 'ultimate'

    # Ult reserve
    ur = ult_reserve_re.search(line)
    if ur:
        current_data['events'].append(f'Ult 예약 SP-{ur.group(2)}')

    # Damage
    dm = dmg_re.search(line)
    if dm:
        crit = ' CRI' if dm.group(5) else ''
        current_data['actions'].append(f'{dm.group(2)} {dm.group(4)}{crit}')

    # DoT
    dt = dot_re.search(line)
    if dt:
        current_data['events'].append(f'{dt.group(1)} {dt.group(2)} -{dt.group(3)}')

    # Kill
    km = kill_re.search(line)
    if km:
        current_data['events'].append(f'KILL: {km.group(1)}')

    # Hard CC
    hc = hard_cc_re.search(line)
    if hc:
        current_data['skill'] = 'CC 행동불가'
        current_data['type'] = 'cc_skip'

    # Trigger
    tr = trigger_re.search(line)
    if tr:
        current_data['events'].append(f'트리거: {tr.group(1)}')

# Last turn
if current_data:
    rows.append(current_data)

# Add victory line from log
for line in log:
    if '🏆' in line or '💀 아군' in line:
        rows.append({
            'turn': '-', 'time': '-', 'extra': False,
            'char': '-', 'side': '-', 'hp': '-',
            'sp_ally': '-', 'sp_enemy': '-',
            'skill': line.strip()[:60], 'type': 'RESULT',
            'actions': [], 'events': []
        })
        break

print(f'Parsed {len(rows)} rows from {len(log)} log lines')

# Load existing xlsx
path = 'docs/M04_vs_M01_배틀시나리오_v2.xlsx'
wb = load_workbook(path)

# Remove old sheet if exists
if '풀 타임라인' in wb.sheetnames:
    del wb['풀 타임라인']

ws = wb.create_sheet('풀 타임라인', 1)  # Insert as 2nd sheet
ws.sheet_properties.tabColor = '34495E'

# Styles
thin = Side(style='thin')
border = Border(top=thin, bottom=thin, left=thin, right=thin)
header_font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='34495E')
data_font = Font(name='Arial', size=9)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)
kill_fill = PatternFill('solid', fgColor='FFD9D9')
ult_fill = PatternFill('solid', fgColor='E8F5E9')
m04_fill = PatternFill('solid', fgColor='F2E6FF')
m01_fill = PatternFill('solid', fgColor='FFF2E6')
extra_fill = PatternFill('solid', fgColor='E3F2FD')
round_fill = PatternFill('solid', fgColor='FFF9C4')
cc_fill = PatternFill('solid', fgColor='EEEEEE')

# Title
ws.merge_cells('A1:J1')
ws['A1'] = f'풀 타임라인 - 전체 {engine.turn_count}턴 상세 로그 (seed=42)'
ws['A1'].font = Font(name='Arial', bold=True, size=13, color='34495E')
ws['A1'].alignment = Alignment(horizontal='center')

# Headers
headers = ['턴', '시간', 'Extra', '캐릭터', '진영', 'HP', 'SP(M04)', 'SP(M01)', '스킬 [타입]', '상세 (피해/이벤트)']
for i, h in enumerate(headers, 1):
    cell = ws.cell(row=3, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border

# Data rows
for r_idx, row in enumerate(rows, 4):
    turn = row['turn']
    time = row['time']
    extra = 'Extra' if row.get('extra') else ''
    char = row['char']
    side = row['side']
    hp = row['hp']
    sp_a = row['sp_ally']
    sp_e = row['sp_enemy']
    skill = row['skill']
    stype = row.get('type', '')
    if stype and stype not in ('ROUND', 'RESULT'):
        skill = f'{skill} [{stype}]'

    actions = row.get('actions', [])
    events = row.get('events', [])
    detail_parts = actions + events
    detail = ' | '.join(detail_parts) if detail_parts else ''

    values = [turn, time, extra, char, side, hp, sp_a, sp_e, skill, detail]
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=r_idx, column=c, value=v)
        cell.font = data_font
        cell.alignment = center if c < 10 else left_wrap
        cell.border = border

    # Row coloring
    fill = None
    if stype == 'ROUND':
        fill = round_fill
    elif stype == 'RESULT':
        fill = PatternFill('solid', fgColor='C8E6C9')
    elif any('KILL' in e for e in events):
        fill = kill_fill
    elif stype == 'cc_skip':
        fill = cc_fill
    elif extra:
        fill = extra_fill
    elif stype == 'ultimate':
        fill = ult_fill
    elif side == 'ally':
        fill = m04_fill
    elif side == 'enemy':
        fill = m01_fill

    if fill:
        for c in range(1, 11):
            ws.cell(row=r_idx, column=c).fill = fill

# Column widths
widths = [5, 6, 6, 10, 6, 12, 8, 8, 22, 50]
for c in range(1, 11):
    ws.column_dimensions[get_column_letter(c)].width = widths[c - 1]

# Freeze panes
ws.freeze_panes = 'A4'

# Legend at bottom
legend_row = len(rows) + 5
ws.merge_cells(f'A{legend_row}:J{legend_row}')
ws.cell(row=legend_row, column=1, value='색상 범례: 보라=M04턴 | 주황=M01턴 | 파랑=Extra Turn(얼티밋) | 초록=얼티밋 발동 | 빨강=킬 발생 | 노랑=라운드 전환 | 회색=CC 행동불가').font = Font(name='Arial', size=9, italic=True)

wb.save(path)
print(f'Saved {len(rows)} rows to sheet "풀 타임라인" in {path}')
