"""Update Character.xlsx Role/Element/SP to match Notion (source of truth)"""
import openpyxl

XLSX_PATH = r'C:\Ai\BattleAgent\data\Character.xlsx'
ROLE_SP = {'Attacker': 6, 'Healer': 3, 'Magician': 4, 'Defender': 3, 'Supporter': 4}

# (row, name, new_element, new_role)
FIXES = [
    (6,  '다비',      'Fire',   'Attacker'),
    (9,  '브라우니',   'Forest', 'Supporter'),
    (15, '구미호',    'Fire',   'Magician'),
    (18, '바토리',    'Forest', 'Attacker'),
    (20, '모나',      'Light',  'Healer'),
    (24, '바리',      'Water',  'Magician'),
    (25, '그릴라',    'Forest', 'Supporter'),
    (28, '이브',      'Water',  'Attacker'),
    (29, '판',        'Forest', 'Magician'),
    (30, '카라라트리', 'Fire',   'Magician'),
    (32, '모건',      'Fire',   'Defender'),
    (34, '미리암',    'Forest', 'Defender'),
]

wb = openpyxl.load_workbook(XLSX_PATH)
ws = wb['Detail<Child>']

changed = 0
for row, name, new_elem, new_role in FIXES:
    cell_name = ws.cell(row=row, column=2).value
    if cell_name != name:
        print(f"WARNING: Row {row} expected '{name}' but found '{cell_name}' - SKIPPING")
        continue

    old_elem = ws.cell(row=row, column=13).value
    old_role = ws.cell(row=row, column=14).value
    old_maxsp = ws.cell(row=row, column=16).value
    old_usesp = ws.cell(row=row, column=17).value

    new_maxsp = ROLE_SP[new_role]
    new_usesp = ROLE_SP[new_role]

    ws.cell(row=row, column=13).value = new_elem
    ws.cell(row=row, column=14).value = new_role
    ws.cell(row=row, column=16).value = new_maxsp
    ws.cell(row=row, column=17).value = new_usesp

    print(f"Row {row} [{name}]: {old_elem}/{old_role}/SP{old_maxsp}/{old_usesp} -> {new_elem}/{new_role}/SP{new_maxsp}/{new_usesp}")
    changed += 1

wb.save(XLSX_PATH)
print(f"\nDone: {changed} rows updated in Character.xlsx")
